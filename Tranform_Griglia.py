import os
import sys
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import pandas as pd
import re
import threading

# Store all unique Volume TT values per model
volume_tt_set_by_model = defaultdict(set)

def extract_model_from_filename(filename):
    """Extract the first 3 characters from filename as model code"""
    base_name = os.path.basename(filename)
    return base_name[:3] if len(base_name) >= 3 else base_name

def collect_volume_tt(folder_path):
    search_pattern = os.path.join(folder_path, '*.xlsx')
    xlsx_files = glob.glob(search_pattern)
    for file in xlsx_files:
        try:
            wb = load_workbook(filename=file, read_only=False, data_only=True)
            if "Griglia Mondo - Volumi" in wb.sheetnames:
                ws = wb["Griglia Mondo - Volumi"]
            else:
                ws = wb["Grid World - Volume"]
            
            for row_dim in ws.row_dimensions.values():
                row_dim.hidden = False
            for col_dim in ws.column_dimensions.values():
                col_dim.hidden = False
            
            col_idx = 4
            model_code = str(ws.cell(row=12, column=4).value or "")
            
            while True:
                volume_head = ws.cell(row=15, column=col_idx).value
                if volume_head is None:
                    break
                col_idx += 1
            
            volume_tt = ws.cell(row=15, column=col_idx - 1).value
            if volume_tt:
                volume_tt_set_by_model[model_code].add(int(volume_tt))
        except Exception as e:
            print(f"Error reading TT in {file}: {e}")

def transform_and_expand(folder_path, De_Para_file_path, output_file_path):
    volume_tt_set_by_model.clear()  # Clear from previous runs
    collect_volume_tt(folder_path)
    volume_tt_sum_by_model = {
        model: sum(tts) for model, tts in volume_tt_set_by_model.items()
    }
    
    search_pattern = os.path.join(folder_path, '*.xlsx')
    xlsx_files = glob.glob(search_pattern)
    os.makedirs(os.path.dirname(output_file_path), exist_ok=True)
    
    merged_rows = []
    header = [
        "Packet", "Code", "Concat", "Version", "Volume Head", "Volume", "Volume TT",
        "SINCOM", "Model code", "Model", "Concat", "Multivalues", "File_Model"
    ]
    
    for file in xlsx_files:
        try:
            # Extract model from filename (first 3 characters)
            file_model = extract_model_from_filename(file)
            print(f"Processing file: {os.path.basename(file)}, extracted model: {file_model}")
            
            wb = load_workbook(filename=file, read_only=False, data_only=True)
            if "Griglia Mondo - Volumi" in wb.sheetnames:
                ws = wb["Griglia Mondo - Volumi"]
            else:
                ws = wb["Grid World - Volume"]
            
            for row_dim in ws.row_dimensions.values():
                row_dim.hidden = False
            for col_dim in ws.column_dimensions.values():
                col_dim.hidden = False
            
            model_metadata = []
            col_idx = 4
            model_code = str(ws.cell(row=12, column=4).value or "")
            model = str(ws.cell(row=11, column=4).value or "")
            volume_tt_sum = volume_tt_sum_by_model.get(model_code, 0)
            
            while True:
                volume_head = ws.cell(row=15, column=col_idx).value
                if volume_head is None:
                    break
                
                version = ws.cell(row=13, column=col_idx).value
                sincom = ws.cell(row=14, column=col_idx).value
                
                metadata = {
                    "col_idx": col_idx,
                    "version": str(version or ""),
                    "volume_head": str(volume_head or ""),
                    "sincom": str(sincom or ""),
                    "model_code": model_code,
                    "model": model
                }
                model_metadata.append(metadata)
                col_idx += 1
            
            for row in range(11, 10001):
                if row == 12 or row == 13 or row == 14 or row == 15:
                    continue
                
                packet = ws.cell(row=row, column=1).value
                code = ws.cell(row=row, column=2).value
                
                if packet is None and code is None:
                    continue
                
                packet_str = str(packet or "")
                code_str = str(code or "").strip()
                concat = str(packet_str + code_str).strip()
                
                for meta in model_metadata:
                    col = meta["col_idx"]
                    volume_val = ws.cell(row=row, column=col).value
                    volume_str = str(volume_val or "")
                    engine_family = str(volume_val).strip() if volume_val else ""
                    
                    merged_rows.append([
                        packet_str,
                        code_str,
                        concat,
                        meta["version"],
                        meta["volume_head"],
                        volume_str,
                        str(volume_tt_sum),
                        meta["sincom"],
                        meta["model_code"],
                        meta["model"],
                        concat,
                        engine_family,
                        file_model  # Add the extracted file model
                    ])
            
            print(f"Processed: {file}")
        except Exception as e:
            print(f"Error processing {file}: {e}")
    
    # Create DataFrame from merged_rows for use in Update_De_Para
    df = pd.DataFrame(merged_rows, columns=header)
    
    # Save the expanded file
    final_wb = Workbook()
    final_ws = final_wb.active
    final_ws.title = "Expanded_Mapped"
    final_ws.append(header)
    for row in merged_rows:
        final_ws.append(row)
    final_wb.save(output_file_path)
    
    # Update De_Para file with the DataFrame
    if De_Para_file_path and os.path.exists(De_Para_file_path):
        Update_De_Para(De_Para_file_path, df)
        
    else:
        print("De_Para file not selected or doesn't exist. Skipping De_Para update.")
    
    print(f"Final file saved to: {output_file_path}")



def Update_De_Para(de_para_file_path, df):
    """
    Finds new unique multivalues, builds their corresponding records, and
    updates the 'tb_de_para' sheet in the specified Excel file.
    
    This is the complete, corrected workflow with pre-filtering and a final check.
    """
    try:
        # --- Step 1: Load All Necessary Data ---
        de_para_mapping = pd.read_excel(de_para_file_path, sheet_name="Coded")
        de_para_mapping = de_para_mapping.drop_duplicates(subset=['Griglia Italiano', 'Griglia Inglês'])
        de_para_mapping = de_para_mapping.dropna(subset=['Griglia Italiano', 'Griglia Inglês'], how='all')

        try:
            # Correctly loading the target sheet to be updated
            tb_de_para = pd.read_excel(de_para_file_path, sheet_name="Coded")
        except ValueError:
            print("Sheet 'tb_de_para' not found. A new one will be created.")
            tb_de_para = pd.DataFrame(columns=['MultiValues', 'Griglia Italiano', 'Griglia Inglês', 'Model', 'Resp.1', 'Resp.2'])

        # --- Step 2: Pre-filter input 'df' based on valid 'Griglia' values ---
        print("Pre-filtering input data against 'Coded' sheet...")
        valid_ita = de_para_mapping['Griglia Italiano'].dropna().astype(str).str.lower().str.strip()
        valid_eng = de_para_mapping['Griglia Inglês'].dropna().astype(str).str.lower().str.strip()
        valid_packets = set(valid_ita) | set(valid_eng)

        df['packet_norm_for_filter'] = df['Packet'].astype(str).str.lower().str.strip()
        
        original_row_count = len(df)
        df = df[df['packet_norm_for_filter'].isin(valid_packets)].copy()
        df.drop(columns=['packet_norm_for_filter'], inplace=True)

        print(f"SUCCESS. Kept {len(df)} of {original_row_count} rows that match a 'Griglia' value.")

        if df.empty:
            print("No rows remained after pre-filtering. No updates to perform.")
            return tb_de_para

        # --- Step 3: Identify New Multivalues from the pre-filtered data ---
        resp1_clean = tb_de_para['Resp.1'].dropna().astype(str).str.lower().str.strip()
        resp2_clean = tb_de_para['Resp.2'].dropna().astype(str).str.lower().str.strip()
        existing_values = set(resp1_clean) | set(resp2_clean)
        

        df['multivalues_clean'] = df['Multivalues'].dropna().astype(str).str.lower().str.strip()
        
        rows_with_new_data = df[~df['multivalues_clean'].isin(existing_values)].copy()
        
        rows_with_new_data.drop(columns=['multivalues_clean'], inplace=True)
        df.drop(columns=['multivalues_clean'], inplace=True)

        if rows_with_new_data.empty:
            print("SUCCESS. No new values were found to add.")
            return tb_de_para

        print(f"SUCCESS. Identified {len(rows_with_new_data['Multivalues'].unique())} new unique values to process and add.")

        # --- Step 4: Build New Records by Matching with 'Coded' Sheet ---
        rows_with_new_data['packet_norm'] = rows_with_new_data['Packet'].str.strip().str.lower()
        de_para_mapping['ita_norm'] = de_para_mapping['Griglia Italiano'].str.strip().str.lower()
        de_para_mapping['eng_norm'] = de_para_mapping['Griglia Inglês'].str.strip().str.lower()

        merged_ita = pd.merge(rows_with_new_data, de_para_mapping, left_on='packet_norm', right_on='ita_norm', how='inner')
        merged_ita['Resp.1'] = merged_ita['Multivalues']
        merged_ita['Resp.2'] = pd.NA

        merged_eng = pd.merge(rows_with_new_data, de_para_mapping, left_on='packet_norm', right_on='eng_norm', how='inner')
        merged_eng['Resp.1'] = pd.NA
        merged_eng['Resp.2'] = merged_eng['Multivalues']

        combined_new = pd.concat([merged_ita, merged_eng], ignore_index=True)

        if combined_new.empty:
            print("New values found, but they did not match any entry in the 'Coded' sheet.")
            return tb_de_para

        new_records = combined_new.groupby(
            ['MultiValues', 'Griglia Italiano', 'Griglia Inglês', 'File_Model'], as_index=False
        ).agg({
            'Resp.1': 'first', 'Resp.2': 'first'
        }).fillna('')

        new_records = new_records.rename(columns={'File_Model': 'Model'})
        # Clean up whitespace and normalize nulls
        new_records['Resp.1'] = new_records['Resp.1'].astype(str).str.strip()
        new_records['Resp.2'] = new_records['Resp.2'].astype(str).str.strip()

        # Filter: Keep only rows where either Resp.1 or Resp.2 is not empty
        filtered_records = new_records[
            (new_records['Resp.1'] != '') |
            (new_records['Resp.2'] != '')
        ]

        final_new_records = filtered_records[tb_de_para.columns].drop_duplicates(subset=['MultiValues'])
       

        print(f"Generated {len(final_new_records)} records before final check.")

        # --- Step 5: Final check to remove values that already exist ---
        # This ensures a MultiValue isn't added if it's already a Resp.1 or Resp.2 value anywhere.
        def is_value_existing(value):
            return any(ev in value for ev in existing_values)

        # Step 3: Apply it to the cleaned column
        if not final_new_records.empty:
            final_new_records['multivalues_clean_final'] = final_new_records['Resp.1'].astype(str).str.lower().str.strip()

            original_count = len(final_new_records)

            # Keep only rows where the value is NOT a substring match with any existing value
            final_new_records = final_new_records[~final_new_records['multivalues_clean_final'].apply(is_value_existing)].copy()

            final_new_records.drop(columns=['multivalues_clean_final'], inplace=True)
            print(final_new_records)
            print(f"After final check, {len(final_new_records)} of {original_count} records are truly new.")
                
        if final_new_records.empty:
            print("All generated records were found to already exist. No new rows will be added.")
            return tb_de_para
            
        print(f"Adding {len(final_new_records)} truly new records to the sheet.")

        # --- Step 6: Append Final New Records and SAVE to Excel ---
        updated_tb_de_para = final_new_records.drop(columns=['Resp.2'])

        try:
            with pd.ExcelWriter(de_para_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                updated_tb_de_para.to_excel(writer, sheet_name='tb_de_para', index=False)
            print(f"✅ Successfully updated '{de_para_file_path}' with {len(final_new_records)} new rows.")
        except Exception as e:
            output_path = os.path.splitext(de_para_file_path)[0] + "_updated_tb_de_para.xlsx"
            updated_tb_de_para.to_excel(output_path, sheet_name='tb_de_para', index=False)
            print(f"Could not write to the original file (Error: {e}). Results saved to: {output_path}")

        return updated_tb_de_para

    except FileNotFoundError:
        print(f"Error: The file '{de_para_file_path}' was not found.")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None


# === Resource path function ===
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller .exe"""
    try:
        base_path = sys._MEIPASS  # Set by PyInstaller
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# === GUI Section ===
root = tk.Tk()
root.title("Griglia Transformer")

# Load Stellantis logo image and display it
stellantis_logo_path = resource_path("assets/Vlc_img.png")
try:
    img_stellantis_logo = Image.open(stellantis_logo_path)
    img_stellantis_logo = img_stellantis_logo.resize((530, 40), Image.Resampling.LANCZOS)
    photo_img_stellantis_logo = ImageTk.PhotoImage(img_stellantis_logo)
    root.image = photo_img_stellantis_logo  # Keep reference

    image_frame = tk.Frame(root, bg="#f0f0f0")
    image_frame.pack(pady=10)
    tk.Label(image_frame, image=photo_img_stellantis_logo, bg="#f0f0f0").pack(side="left", padx=20)
except Exception as e:
    print(f"Could not load logo: {e}")

# Input folder selection
tk.Label(root, text="Select Input Folder:").pack(anchor="w", padx=10)
input_folder_var = tk.StringVar()
tk.Entry(root, textvariable=input_folder_var, width=70).pack(padx=10, pady=2)
tk.Button(root, text="Browse Input Folder", command=lambda: input_folder_var.set(filedialog.askdirectory())).pack(pady=5)

# De_Para file selection
tk.Label(root, text="Select De_Para File (optional):").pack(anchor="w", padx=10)
de_para_file_var = tk.StringVar()
tk.Entry(root, textvariable=de_para_file_var, width=70).pack(padx=10, pady=2)
tk.Button(root, text="Browse De_Para File", 
          command=lambda: de_para_file_var.set(filedialog.askopenfilename(
              title="Select De_Para File",
              filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
          ))).pack(pady=5)

# Output folder selection
tk.Label(root, text="Select Output Folder:").pack(anchor="w", padx=10)
output_folder_var = tk.StringVar()
tk.Entry(root, textvariable=output_folder_var, width=70).pack(padx=10, pady=2)
tk.Button(root, text="Browse Output Folder", command=lambda: output_folder_var.set(filedialog.askdirectory())).pack(pady=5)

# Progress bar and label
progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
progress_bar.pack(pady=10)
progress_label = tk.Label(root, text="")
progress_label.pack()

def threaded_processing(input_folder, de_para_file, output_folder):
    try:
        progress_bar.start(10)  # Start progress animation
        progress_label.config(text="Processing...")
        
        output_file_path = os.path.join(output_folder, "Expanded_Mapped_File.xlsx")
        
        transform_and_expand(input_folder, de_para_file, output_file_path)
        
        progress_label.config(text="Done!")
        messagebox.showinfo("Success", f"File saved at:\n{output_file_path}")
        progress_bar.stop()
    except Exception as e:
        messagebox.showerror("Processing Error", str(e))
        print(f"Processing error: {e}")
    finally:
        progress_bar.stop()
        progress_label.config(text="")
        run_button.config(state="normal")

def run_processing():
    input_folder = input_folder_var.get()
    output_folder = output_folder_var.get()
    de_para_file = de_para_file_var.get()
    
    if not input_folder or not output_folder:
        messagebox.showerror("Error", "Please select both input and output folders.")
        return
    
    # Validate that input folder exists and contains Excel files
    if not os.path.exists(input_folder):
        messagebox.showerror("Error", "Input folder does not exist.")
        return
    
    # Check if there are any Excel files in the input folder
    search_pattern = os.path.join(input_folder, '*.xlsx')
    xlsx_files = glob.glob(search_pattern)
    if not xlsx_files:
        messagebox.showerror("Error", "No Excel files found in the input folder.")
        return
    
    # Validate De_Para file if provided
    if de_para_file and not os.path.exists(de_para_file):
        messagebox.showerror("Error", "Selected De_Para file does not exist.")
        return
    
    # Validate that De_Para file has the required sheet if provided
    if de_para_file:
        try:
            wb = pd.ExcelFile(de_para_file)
            if "Coded" not in wb.sheet_names:
                messagebox.showerror("Error", "De_Para file must contain a 'Coded' sheet.")
                return
        except Exception as e:
            messagebox.showerror("Error", f"Error reading De_Para file: {e}")
            return
    
    # Create output directory if it doesn't exist
    try:
        os.makedirs(output_folder, exist_ok=True)
    except Exception as e:
        messagebox.showerror("Error", f"Cannot create output folder: {e}")
        return
    
    run_button.config(state="disabled")  # Disable button to prevent re-click
    threading.Thread(target=threaded_processing, args=(input_folder, de_para_file, output_folder), daemon=True).start()

run_button = tk.Button(root, text="Run Processing", command=run_processing, height=2, width=30, 
                      bg="#4CAF50", fg="white", font=("Helvetica", 12, "bold"))
run_button.pack(pady=20)

# Add status information
status_frame = tk.Frame(root)
status_frame.pack(pady=10)

tk.Label(status_frame, text="Status Information:", font=("Helvetica", 10, "bold")).pack(anchor="w")
status_text = tk.Text(status_frame, height=8, width=80, wrap=tk.WORD)
status_text.pack(padx=10, pady=5)

# Redirect print statements to the status text widget
class TextRedirector:
    def __init__(self, widget):
        self.widget = widget

    def write(self, str):
        self.widget.insert(tk.END, str)
        self.widget.see(tk.END)
        root.update_idletasks()

    def flush(self):
        pass

# Redirect stdout to the text widget
sys.stdout = TextRedirector(status_text)

# Add a clear button for the status text
def clear_status():
    status_text.delete(1.0, tk.END)

tk.Button(status_frame, text="Clear Status", command=clear_status).pack(pady=5)


# Center the window
root.update_idletasks()
width = root.winfo_width()
height = root.winfo_height()
x = (root.winfo_screenwidth() // 2) - (width // 2)
y = (root.winfo_screenheight() // 2) - (height // 2)
root.geometry(f"{width}x{height}+{x}+{y}")

root.mainloop()
